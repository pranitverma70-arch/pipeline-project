import re
import pdfplumber
import openpyxl
import logging

logger = logging.getLogger(__name__)

# Known pipelines for auto-detection
KNOWN_PIPELINES = {
    "Mundra-Delhi Pipeline (MDPL)": ["Mundra-Delhi Pipeline", "MDPL"],
    "Gujarat-Punjab Pipeline (GPPL)": ["Gujarat-Punjab Pipeline", "GPPL"],
    "Mumbai-Pune Pipeline (MPPL)": ["Mumbai-Pune Pipeline", "MPPL"]
}

def detect_pipeline_in_text(text: str) -> str:
    text_lower = text.lower()
    for full_name, search_terms in KNOWN_PIPELINES.items():
        for term in search_terms:
            if term.lower() in text_lower:
                return full_name
    return None

# List of parameter names we want to extract
PARAMETERS_TO_EXTRACT = [
    "Scrapper Pigging",
    "PIG Residue Analyisis",
    "Corrosion Coupon",
    "Corrosion Probe",
    "PSP Reading at Feeding Points, Casing, Mid Point",
    "PSP ON Potential",
    "PSP Instant Off Potential",
    "Current Consumption Data",
    "Cathodic Protection Rectifiers",
    "Polarization Cells",
    "Crossing Location Data",
    "IJ Health Report",
    "Surge Diverter in IJ",
    "Anode Bed Data",
    "Line Current Data",
    "CIPL",
    "DCVG",
    "Coating Conduction Survey",
    "AC Inerferance  Survey",
    "DC Inerferance  Survey",
    "Soil Resistivity Survey",
    "Static Leak Simulation",
    "Dynamic Leak Simulation",
    "ROU Management",
    "Audit Management"
]

def extract_from_pdf(filepath: str) -> tuple:
    """Extract parameters and scores from a PDF file using pdfplumber and regex."""
    extracted = {}
    text = ""
    try:
        text = ""
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"

        # Regex heuristic: parameter name followed by some text, spaces, or colons, then a number
        for param in PARAMETERS_TO_EXTRACT:
            # We want to match param name exactly (case insensitive) followed by non-digits up to a colon or equals, then a float
            # Examples: "DCVG: 4.5", "DCVG Score 4.36", "DCVG = 4.36"
            escaped_param = re.escape(param).replace("\\ ", "\\s*")
            pattern = re.compile(rf"{escaped_param}[^0-9\n]{{0,20}}?(\d+(?:\.\d+)?)", re.IGNORECASE)
            match = pattern.search(text)
            if match:
                try:
                    val = float(match.group(1))
                    extracted[param] = {"value": val, "confidence": 0.90}
                    logger.info(f"Extracted from PDF: {param} = {val} (conf: 0.90)")
                except ValueError:
                    pass
    except Exception as e:
        logger.error(f"Error extracting from PDF {filepath}: {e}")

    detected_pipeline = detect_pipeline_in_text(text)
    return extracted, detected_pipeline

def extract_from_excel(filepath: str) -> tuple:
    """Extract parameters and scores from an Excel file using openpyxl."""
    extracted = {}
    full_text = ""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                row_str = " ".join([str(c) for c in row if c is not None])
                full_text += row_str + " "
                
                for param in PARAMETERS_TO_EXTRACT:
                    # If param name is in this row (case insensitive check)
                    if param.lower() in row_str.lower():
                        # Try to find a numeric value in the row (after the param name if possible)
                        # We will just look at all cell values in the row and take the first float that looks like a score
                        for cell_value in row:
                            if isinstance(cell_value, (int, float)):
                                extracted[param] = {"value": float(cell_value), "confidence": 0.80}
                                logger.info(f"Extracted from Excel: {param} = {cell_value} (conf: 0.80)")
                                break # Found a score for this parameter in this row
    except Exception as e:
        logger.error(f"Error extracting from Excel {filepath}: {e}")

    detected_pipeline = detect_pipeline_in_text(full_text)
    return extracted, detected_pipeline

def process_upload(filepath: str, filename: str) -> dict:
    """
    Process an uploaded file and extract numeric scores.
    Uses pdfplumber for PDFs and openpyxl for Excel files.
    """
    filename_lower = filename.lower()
    extracted_data = {}
    detected_pipeline = None

    if filename_lower.endswith(".pdf"):
        extracted_data, detected_pipeline = extract_from_pdf(filepath)
    elif filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls"):
        extracted_data, detected_pipeline = extract_from_excel(filepath)
    else:
        logger.warning(f"Unsupported file format for extraction: {filename}")

    return {
        "data": extracted_data,
        "pipeline_name": detected_pipeline
    }
