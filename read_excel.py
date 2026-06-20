import pandas as pd
import sys

try:
    # We use openpyxl to get formulas and pandas to get data if needed.
    import openpyxl
    wb = openpyxl.load_workbook('Pipeline Health index_3.xlsx', data_only=False)
    print("Sheets in workbook:")
    for sheet_name in wb.sheetnames:
        print(f"- {sheet_name}")
        sheet = wb[sheet_name]
        # print first few rows to see structure
        for row in sheet.iter_rows(min_row=1, max_row=10, values_only=False):
            row_data = []
            for cell in row:
                if cell.value is not None:
                    row_data.append(f"{cell.coordinate}: {cell.value}")
            if row_data:
                print("  ", " | ".join(row_data))
        print("---")
except Exception as e:
    print(f"Failed to read with openpyxl: {e}")
