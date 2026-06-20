import openpyxl
wb = openpyxl.load_workbook('Pipeline Health index_3.xlsx', data_only=True)

print('=== Calculation_1 (Main Parameter Scores) ===')
s = wb['Calculation_1']
for row in range(2, 15):
    vals = []
    for col in range(1, 5):
        v = s.cell(row=row, column=col).value
        if v is not None:
            vals.append(str(v))
    if vals:
        print('  Row %d: %s' % (row, ' | '.join(vals)))

print('')
print('=== Calcualtion of Index (Detailed Scores) ===')
s2 = wb['Calcualtion of Index']
for row in range(1, 35):
    a = s2.cell(row=row, column=1).value  # Details
    d = s2.cell(row=row, column=4).value  # Weightage
    e = s2.cell(row=row, column=5).value  # Job Done
    f = s2.cell(row=row, column=6).value  # Data Analysis
    g = s2.cell(row=row, column=7).value  # Compliance
    h = s2.cell(row=row, column=8).value  # Incomplete
    i_val = s2.cell(row=row, column=9).value  # Marks
    if a is not None:
        line = '  Row %d: %s' % (row, str(a))
        if d is not None:
            line += ' | W=%s' % str(d)
        if i_val is not None:
            line += ' | Score=%s' % str(i_val)
        print(line)

print('')
print('=== Input Data (Key Values) ===')
s3 = wb['Input Data ']
for row in range(1, 50):
    vals = []
    for col in range(1, 7):
        v = s3.cell(row=row, column=col).value
        if v is not None:
            vals.append('%s=%s' % (s3.cell(row=row, column=col).coordinate, str(v)))
    if vals:
        print('  Row %d: %s' % (row, ' | '.join(vals)))

print('')
print('=== Soil Resistivity Analysis ===')
s4 = wb['Soil Resisivty Analysis ']
for row in range(1, 10):
    vals = []
    for col in range(1, 7):
        v = s4.cell(row=row, column=col).value
        if v is not None:
            vals.append('%s=%s' % (s4.cell(row=row, column=col).coordinate, str(v)))
    if vals:
        print('  Row %d: %s' % (row, ' | '.join(vals)))

print('')
print('=== Audit & ROU Management ===')
s5 = wb['Audit & ROU Management']
for row in range(1, 10):
    vals = []
    for col in range(1, 7):
        v = s5.cell(row=row, column=col).value
        if v is not None:
            vals.append('%s=%s' % (s5.cell(row=row, column=col).coordinate, str(v)))
    if vals:
        print('  Row %d: %s' % (row, ' | '.join(vals)))
